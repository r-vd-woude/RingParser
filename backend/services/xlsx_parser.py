from openpyxl import load_workbook
from pathlib import Path
from backend.utils.type_inference import _infer_column_types
from typing import Dict, Any
from backend.services.base_parser import BaseParser


class XLSXParser(BaseParser):
    """Parser for XLSX files using openpyxl"""

    def __init__(self):
        pass

    def _parse_file_sync(self, file_path: Path) -> Dict[str, Any]:
        workbook = load_workbook(filename=file_path, read_only=True)
        return self._parse_workbook(workbook, name=file_path.name)

    def _parse_workbook(self, workbook, name: str) -> Dict[str, Any]:
        sheet = workbook.active
        headers = [str(cell.value) if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        from backend.config import MAX_DATA_ROWS
        data_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(data_rows) >= MAX_DATA_ROWS:
                raise ValueError(f"File exceeds the maximum of {MAX_DATA_ROWS:,} data rows")
            data_rows.append(row)

        sample_data = data_rows[:10]
        column_info = _infer_column_types(headers, data_rows)

        return {
            "filename": name,
            "total_rows": len(data_rows),
            "columns": column_info,
            "sample_data": sample_data,
            "data_rows": data_rows,
            "headers": headers,
        }

